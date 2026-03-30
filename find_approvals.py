#!/usr/bin/env python3
"""
Find and fill FDA, TGA, and MedSafe approval dates for Pharmac applications.

This script reads the Pharmac applications from 'Pharmac applications.xlsx',
looks up each drug in the FDA, TGA, and MedSafe approval databases,
and outputs a CSV with the earliest approval dates found for each regulator.

Fuzzy string matching is used to handle name variations between databases.

Usage:
    python find_approvals.py

Output:
    pharmac_approvals.csv  - Pharmac applications with FDA/TGA/MedSafe approval dates
"""

import os
import re
import csv
import warnings
from datetime import datetime

import pandas as pd
import openpyxl
from thefuzz import fuzz

warnings.filterwarnings("ignore")

# --- Paths ---
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
FDA_PRODUCTS = os.path.join(DATA_DIR, "fda", "Products.txt")
FDA_SUBMISSIONS = os.path.join(DATA_DIR, "fda", "Submissions.txt")
MEDSAFE_DB = os.path.join(DATA_DIR, "medsafe", "medsafe_register.csv")
TGA_DB = os.path.join(DATA_DIR, "tga", "tga_artg.csv")
PHARMAC_XLSX = os.path.join(os.path.dirname(__file__), "Pharmac applications.xlsx")
OUTPUT_CSV = os.path.join(os.path.dirname(__file__), "pharmac_approvals.csv")

# --- Fuzzy matching threshold ---
FUZZY_THRESHOLD = 80


def normalize_name(name):
    """Normalize a drug name for fuzzy matching."""
    if not name:
        return ""
    name = str(name).lower()
    # Remove common dosage form suffixes
    name = re.sub(
        r"\b(tablets?|capsules?|injections?|solutions?|syrups?|creams?|ointments?|"
        r"gels?|powders?|drops?|patches?|sprays?|inhalers?|suspensions?|films?|"
        r"coated|oral|topical|intravenous|intramuscular|subcutaneous|"
        r"extended|modified|delayed|sustained|prolonged|immediate|controlled|"
        r"release|long|short|acting)\b",
        " ",
        name,
    )
    # Remove hydrochloride, sodium, etc. salt forms
    name = re.sub(
        r"\b(hydrochloride|hcl|sodium|potassium|acetate|phosphate|sulfate|"
        r"mesylate|maleate|tartrate|citrate|gluconate|succinate|fumarate|"
        r"malonate|lactate|chloride|bromide|iodide|carbonate|bicarbonate|"
        r"monohydrate|dihydrate|trihydrate|anhydrous|alfa|beta|gamma)\b",
        " ",
        name,
    )
    # Remove punctuation except letters and spaces
    name = re.sub(r"[^\w\s]", " ", name)
    # Collapse whitespace
    name = re.sub(r"\s+", " ", name).strip()
    return name


def split_combination_drug(name):
    """Split a combination drug name into individual components."""
    if not name:
        return []
    name = str(name)
    # Split on common combination separators
    parts = re.split(r"[/+]|\band\b|\bwith\b", name, flags=re.I)
    result = []
    for part in parts:
        part = part.strip()
        if len(part) > 2:
            result.append(part)
    return result if result else [name]


def get_primary_ingredient(name):
    """Extract the primary (first) ingredient from a drug name."""
    if not name:
        return ""
    parts = split_combination_drug(name)
    primary = parts[0] if parts else name
    # Remove trailing formulation info (e.g., "200mg", "(oral)")
    primary = re.sub(r"\s+\d.*$", "", primary)
    primary = re.sub(r"\s*\(.*\)$", "", primary)
    primary = re.sub(r"\s*[-–—]\s*(Growth Hormone|oral|injection|topical).*$", "", primary, flags=re.I)
    return primary.strip()


def best_match_score(query, candidates):
    """Return the best fuzzy match score between query and a list of candidates."""
    if not query or not candidates:
        return 0
    query_norm = normalize_name(query)
    best = 0
    for candidate in candidates:
        if not candidate:
            continue
        candidate_norm = normalize_name(str(candidate))
        # Try several fuzzy matching strategies
        scores = [
            fuzz.token_set_ratio(query_norm, candidate_norm),
            fuzz.partial_ratio(query_norm, candidate_norm),
            fuzz.ratio(query_norm, candidate_norm),
        ]
        score = max(scores)
        if score > best:
            best = score
    return best


# =============================================================================
# FDA Database
# =============================================================================

def load_fda_database():
    """Load and merge FDA products and submissions data."""
    if not os.path.exists(FDA_PRODUCTS) or not os.path.exists(FDA_SUBMISSIONS):
        print("WARNING: FDA database files not found. Run download_databases.py first.")
        return None

    print("Loading FDA database...")
    prods = pd.read_csv(FDA_PRODUCTS, sep="\t", dtype=str, encoding="latin-1")
    subs = pd.read_csv(FDA_SUBMISSIONS, sep="\t", dtype=str, encoding="latin-1")

    # Get earliest original approval date per application
    orig = subs[
        (subs["SubmissionType"] == "ORIG") & (subs["SubmissionStatus"] == "AP")
    ][["ApplNo", "SubmissionStatusDate"]].copy()
    orig["ApprovalDate"] = pd.to_datetime(
        orig["SubmissionStatusDate"], errors="coerce"
    )
    earliest = orig.groupby("ApplNo")["ApprovalDate"].min().reset_index()
    earliest.rename(columns={"ApprovalDate": "FDA_ApprovalDate"}, inplace=True)

    # Merge with products
    merged = prods.merge(earliest, on="ApplNo", how="left")
    merged = merged.dropna(subset=["FDA_ApprovalDate"])
    merged["DrugName_norm"] = merged["DrugName"].str.lower().str.strip()
    merged["Ingredient_norm"] = merged["ActiveIngredient"].str.lower().str.strip()

    print(f"FDA database: {len(merged)} products with approval dates")
    return merged


def lookup_fda(fda_db, chemical_name, brand_name):
    """Look up FDA approval date for a drug.

    Returns the earliest FDA approval date, or None if not found.
    """
    if fda_db is None:
        return None

    candidates = []

    # Extract individual ingredients
    ingredients = split_combination_drug(chemical_name)
    primary = get_primary_ingredient(chemical_name)

    # Build a list of search terms from general to specific
    search_terms = []
    for ing in ingredients + [primary]:
        ing = ing.strip()
        if not ing:
            continue
        search_terms.append(ing.lower())
        # Also add first 1-2 words for partial matching
        words = ing.lower().split()
        if len(words) >= 2:
            search_terms.append(" ".join(words[:2]))
        if len(words) >= 1:
            search_terms.append(words[0])

    # Remove duplicates while preserving order
    seen_terms = set()
    unique_terms = []
    for t in search_terms:
        if t not in seen_terms and len(t) >= 4:
            seen_terms.add(t)
            unique_terms.append(t)

    # Search strategies (in order of preference):
    # 1. Substring match on ingredient
    for term in unique_terms:
        matches = fda_db[fda_db["Ingredient_norm"].str.contains(term, na=False, regex=False)]
        if not matches.empty:
            candidates.extend(matches["FDA_ApprovalDate"].dropna().tolist())

    # 2. Exact match on drug name (brand name)
    if brand_name:
        brand_lower = brand_name.lower().strip()
        matches = fda_db[fda_db["DrugName_norm"].str.contains(brand_lower, na=False, regex=False)]
        if not matches.empty:
            candidates.extend(matches["FDA_ApprovalDate"].dropna().tolist())

    # 3. Fuzzy match on ingredient (for spelling variations)
    if not candidates:
        primary_norm = normalize_name(primary)
        if len(primary_norm) >= 4:
            # Only check against unique ingredients for efficiency
            unique_ingredients = fda_db.drop_duplicates(subset=["Ingredient_norm"])
            for _, row in unique_ingredients.iterrows():
                row_norm = normalize_name(str(row.get("Ingredient_norm", "")))
                score = fuzz.token_set_ratio(primary_norm, row_norm)
                if score >= FUZZY_THRESHOLD:
                    # Get all products with this ingredient
                    ing_matches = fda_db[fda_db["Ingredient_norm"] == row["Ingredient_norm"]]
                    candidates.extend(ing_matches["FDA_ApprovalDate"].dropna().tolist())

    if candidates:
        return min(candidates)
    return None


# =============================================================================
# MedSafe Database
# =============================================================================

def load_medsafe_database():
    """Load MedSafe register database."""
    if not os.path.exists(MEDSAFE_DB):
        print("WARNING: MedSafe database not found. Run download_databases.py first.")
        return None

    print("Loading MedSafe database...")
    df = pd.read_csv(MEDSAFE_DB, dtype=str, encoding="utf-8")
    df.columns = [c.strip() for c in df.columns]
    df["Ingredient_norm"] = df["Active ingredients"].str.lower().str.strip()

    # Parse approval dates
    def parse_nz_date(date_str):
        if pd.isna(date_str) or not str(date_str).strip():
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %b %Y"):
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except ValueError:
                continue
        return None

    df["ApprovalDate"] = df["Approval date"].apply(parse_nz_date)
    print(f"MedSafe database: {len(df)} products")
    return df


def lookup_medsafe(medsafe_db, chemical_name, brand_name):
    """Look up MedSafe approval date for a drug.

    Returns the earliest MedSafe approval date, or None if not found.
    """
    if medsafe_db is None:
        return None

    candidates = []
    ingredients = split_combination_drug(chemical_name)
    primary = get_primary_ingredient(chemical_name)

    # Search strategies:
    # 1. Exact substring match on ingredients
    for ing in ingredients + [primary]:
        ing_lower = ing.lower().strip()
        if len(ing_lower) < 3:
            continue
        matches = medsafe_db[
            medsafe_db["Ingredient_norm"].str.contains(ing_lower, na=False, regex=False)
        ]
        if not matches.empty:
            dates = matches["ApprovalDate"].dropna().tolist()
            candidates.extend(dates)

    # 2. Brand name match on Product column
    if brand_name and not candidates:
        brand_lower = brand_name.lower().strip()
        matches = medsafe_db[
            medsafe_db["Product"].str.lower().str.contains(brand_lower, na=False, regex=False)
        ]
        if not matches.empty:
            dates = matches["ApprovalDate"].dropna().tolist()
            candidates.extend(dates)

    # 3. Fuzzy match on ingredients (for cases with spelling variations)
    if not candidates:
        for ing in ingredients + [primary]:
            ing_norm = normalize_name(ing)
            if len(ing_norm) < 4:
                continue
            for _, row in medsafe_db.iterrows():
                row_norm = normalize_name(str(row.get("Ingredient_norm", "")))
                score = fuzz.token_set_ratio(ing_norm, row_norm)
                if score >= FUZZY_THRESHOLD:
                    if row.get("ApprovalDate"):
                        candidates.append(row["ApprovalDate"])

    if candidates:
        return min(candidates)
    return None


# =============================================================================
# TGA Database
# =============================================================================

def load_tga_database():
    """Load TGA ARTG database."""
    if not os.path.exists(TGA_DB):
        print("WARNING: TGA database not found. Run download_databases.py first.")
        return None

    print("Loading TGA database...")
    df = pd.read_csv(TGA_DB, dtype=str, encoding="utf-8")
    df.columns = [c.strip() for c in df.columns]

    # Normalise column names: Selenium scraper uses ARTG_ID / Name / RegistrationDate;
    # a manually downloaded ARTG extract may use different names.
    col_map = {}
    for col in df.columns:
        low = col.lower().strip()
        if low in ("name", "product description", "artg description", "description"):
            col_map[col] = "Name"
        elif low in ("registrationdate", "registration date", "artg start date",
                     "start date", "approval date"):
            col_map[col] = "RegistrationDate"
    if col_map:
        df = df.rename(columns=col_map)

    if "Name" in df.columns:
        df["Name_norm"] = df["Name"].str.lower().str.strip()
    if "RegistrationDate" in df.columns:
        df["_date"] = pd.to_datetime(df["RegistrationDate"], errors="coerce")

    print(f"TGA database: {len(df)} products")
    return df


def lookup_tga(tga_db, chemical_name, brand_name):
    """Look up TGA registration date for a drug.

    Returns the earliest TGA registration date, or None if not found.
    """
    if tga_db is None:
        return None
    if "Name_norm" not in tga_db.columns or "_date" not in tga_db.columns:
        return None

    candidates = []
    ingredients = split_combination_drug(chemical_name)
    primary = get_primary_ingredient(chemical_name)

    # 1. Substring match of each ingredient against Name
    for ing in ingredients + [primary]:
        ing_lower = ing.lower().strip()
        if len(ing_lower) < 3:
            continue
        matches = tga_db[
            tga_db["Name_norm"].str.contains(ing_lower, na=False, regex=False)
        ]
        if not matches.empty:
            candidates.extend(matches["_date"].dropna().tolist())

    # 2. Brand name substring match
    if brand_name and not candidates:
        brand_lower = brand_name.lower().strip()
        if len(brand_lower) >= 3:
            matches = tga_db[
                tga_db["Name_norm"].str.contains(brand_lower, na=False, regex=False)
            ]
            if not matches.empty:
                candidates.extend(matches["_date"].dropna().tolist())

    # 3. Fuzzy match on primary ingredient against Name
    if not candidates:
        primary_norm = normalize_name(primary)
        if len(primary_norm) >= 4:
            for _, row in tga_db.iterrows():
                row_norm = normalize_name(str(row.get("Name_norm", "")))
                if fuzz.token_set_ratio(primary_norm, row_norm) >= FUZZY_THRESHOLD:
                    if pd.notna(row.get("_date")):
                        candidates.append(row["_date"])

    if candidates:
        return min(candidates)
    return None


# =============================================================================
# Main Processing
# =============================================================================

def load_pharmac_applications():
    """Load Pharmac applications from Excel file."""
    print("Loading Pharmac applications...")
    wb = openpyxl.load_workbook(PHARMAC_XLSX)
    ws = wb["Sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    print(f"Loaded {len(rows)} Pharmac applications")
    return headers, rows


def main():
    # Load all databases
    fda_db = load_fda_database()
    medsafe_db = load_medsafe_database()
    tga_db = load_tga_database()

    if fda_db is None and medsafe_db is None and tga_db is None:
        print("ERROR: No databases found. Please run download_databases.py first.")
        return

    # Load Pharmac applications
    headers, rows = load_pharmac_applications()

    # Add approval date columns
    new_cols = ["FDA_ApprovalDate", "TGA_ApprovalDate", "MedSafe_ApprovalDate"]
    output_headers = headers + new_cols

    print(f"\nProcessing {len(rows)} applications...")
    results = []
    found_fda = 0
    found_tga = 0
    found_medsafe = 0

    for i, row in enumerate(rows):
        if i % 100 == 0:
            print(f"  Processing row {i+1}/{len(rows)}...")

        chemical_name = str(row.get("Chemical_Name__c", "") or "")
        brand_name = str(row.get("Brand_Name__c", "") or "")

        # Lookup FDA approval date
        fda_date = lookup_fda(fda_db, chemical_name, brand_name)
        if fda_date:
            found_fda += 1

        # Lookup TGA approval date
        tga_date = lookup_tga(tga_db, chemical_name, brand_name)
        if tga_date:
            found_tga += 1

        # Lookup MedSafe approval date
        medsafe_date = lookup_medsafe(medsafe_db, chemical_name, brand_name)
        if medsafe_date:
            found_medsafe += 1

        # Format dates as strings
        def fmt_date(d):
            if d is None:
                return ""
            if hasattr(d, "strftime"):
                return d.strftime("%Y-%m-%d")
            return str(d)[:10]

        result = dict(row)
        result["FDA_ApprovalDate"] = fmt_date(fda_date)
        result["TGA_ApprovalDate"] = fmt_date(tga_date)
        result["MedSafe_ApprovalDate"] = fmt_date(medsafe_date)
        results.append(result)

    # Write output CSV
    print(f"\nWriting results to {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=output_headers)
        writer.writeheader()
        writer.writerows(results)

    print(f"\nDone! Results saved to {OUTPUT_CSV}")
    print(f"  FDA approval dates found: {found_fda}/{len(rows)} ({100*found_fda//len(rows)}%)")
    print(f"  TGA approval dates found: {found_tga}/{len(rows)} ({100*found_tga//len(rows)}%)")
    print(f"  MedSafe approval dates found: {found_medsafe}/{len(rows)} ({100*found_medsafe//len(rows)}%)")


if __name__ == "__main__":
    main()
